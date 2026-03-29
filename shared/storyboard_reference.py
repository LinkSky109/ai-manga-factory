from __future__ import annotations

from collections import Counter, defaultdict
import json
from pathlib import Path
from typing import Any

from backend.config import REFERENCE_DATA_DIR

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime optional dependency
    load_workbook = None


REFERENCE_DIR = REFERENCE_DATA_DIR
REFERENCE_JSON = REFERENCE_DIR / "storyboard_reference_ep1.json"
PROFILE_JSON = REFERENCE_DIR / "storyboard_profile.json"

DEFAULT_GROUP_BEATS = ["开场钩子", "关系建立", "冲突升级", "高潮前停顿", "高潮", "尾钩"]
DEFAULT_GROUP_STYLE_BLOCKS = [
    {
        "group": 1,
        "beat": "开场钩子",
        "focus": "环境建场、群像期待、目标人物未出场先造势",
        "size_candidates": ["大全景", "中远景", "中景", "近景"],
        "movement_candidates": ["航拍俯拍+缓慢下摇", "横向平移掠镜", "微推", "定镜"],
        "dialogue_density": "medium",
        "audio_density": "high",
        "audio_focus": "环境声先行，先造世界再给人物",
    },
    {
        "group": 2,
        "beat": "关系建立",
        "focus": "人物关系、信息补充、视角切换、把目标和处境交代清楚",
        "size_candidates": ["中近景", "近景", "中景", "主观远景"],
        "movement_candidates": ["定镜", "快推", "起身跟拍", "穿雨微推"],
        "dialogue_density": "high",
        "audio_density": "high",
        "audio_focus": "对白和环境音并存，建立叙事基线",
    },
    {
        "group": 3,
        "beat": "冲突升级",
        "focus": "正反打、规则解释、局势升级、把观众推进冲突中心",
        "size_candidates": ["双人中景", "中近景", "近特写", "近景"],
        "movement_candidates": ["微推", "定镜", "突推", "望向反打"],
        "dialogue_density": "high",
        "audio_density": "medium",
        "audio_focus": "对白驱动，辅以重点音效",
    },
    {
        "group": 4,
        "beat": "高潮前停顿",
        "focus": "情绪压缩、关键判断、爆发前的短暂停顿",
        "size_candidates": ["全景", "群像中景", "正反打特写", "特写"],
        "movement_candidates": ["突切+抬镜", "环绕半圈", "缓推到极近", "锁定"],
        "dialogue_density": "medium",
        "audio_density": "medium",
        "audio_focus": "收窄环境，给情绪留压强",
    },
    {
        "group": 5,
        "beat": "高潮",
        "focus": "关键场面爆发、情绪兑现、名台词或动作爆点落地",
        "size_candidates": ["近景", "中近景", "特写", "双人近景"],
        "movement_candidates": ["慢推", "正拍微推", "反打定格", "插入"],
        "dialogue_density": "high",
        "audio_density": "high",
        "audio_focus": "动作音效、配乐和对白共同推高潮",
    },
    {
        "group": 6,
        "beat": "尾钩",
        "focus": "余波、反应、悬念停顿、下一章钩子",
        "size_candidates": ["中景", "近特写", "主观镜头", "插入特写"],
        "movement_candidates": ["环境停顿式定镜", "声音先行+快速切", "甩镜跟拍", "定住冲击画面"],
        "dialogue_density": "medium",
        "audio_density": "high",
        "audio_focus": "先给尾音，再给信息钩子",
    },
]

DEFAULT_STORYBOARD_PROFILE = {
    "source": "内置默认模板",
    "target_duration_seconds": 96.0,
    "target_duration_range": [90.0, 120.0],
    "group_count": 6,
    "group_durations": [16.5, 16.6, 15.6, 15.5, 17.0, 14.5],
    "group_beats": DEFAULT_GROUP_BEATS,
    "group_style_blocks": DEFAULT_GROUP_STYLE_BLOCKS,
    "default_shot_count": 10,
    "default_keyframe_count": 4,
    "required_fields": [
        "分组",
        "15秒段",
        "镜头号",
        "时长(s)",
        "起始时间",
        "结束时间",
        "场景/时间",
        "镜头景别",
        "镜头运动",
        "画面内容",
        "人物动作/神态",
        "台词对白",
        "角色",
        "音效",
    ],
    "chapter_requirements": {
        "storyboard_files": ["storyboard.json", "storyboard.csv", "storyboard.xlsx"],
        "video_files": ["preview/chapter_preview.mp4", "delivery/chapter_final_cut.mp4"],
        "qa_files": ["qa/qa_report.md", "qa/qa_snapshot.json"],
        "audio_files": ["audio/audio_plan.json", "audio/voiceover.mp3", "audio/ambience.wav"],
    },
    "notes": [
        "单章成片总时长控制在 90-120 秒。",
        "必须保留分组、镜头时长、景别、运镜、对白、角色和音效字段。",
        "允许关键帧少于镜头数，但不允许章节没有完整分镜表和章节视频。",
    ],
}


def load_storyboard_profile() -> dict[str, Any]:
    if PROFILE_JSON.exists():
        profile = json.loads(PROFILE_JSON.read_text(encoding="utf-8"))
        reference_payload = json.loads(REFERENCE_JSON.read_text(encoding="utf-8")) if REFERENCE_JSON.exists() else None
        return _normalize_storyboard_profile(profile, reference_payload=reference_payload)
    return _normalize_storyboard_profile(DEFAULT_STORYBOARD_PROFILE.copy())


def save_storyboard_reference_from_workbook(workbook_path: Path) -> dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl，无法导入 Excel 分镜参考")
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    workbook = load_workbook(workbook_path, data_only=True)
    sheets: list[dict[str, Any]] = []
    profile = DEFAULT_STORYBOARD_PROFILE.copy()

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        normalized_rows = []
        for row in rows:
            normalized_rows.append([_normalize_cell(cell) for cell in row])

        header_row = normalized_rows[2] if len(normalized_rows) >= 3 else []
        data_rows = []
        for row in normalized_rows[3:]:
            if not any(cell not in ("", None) for cell in row):
                continue
            record = {}
            for index, header in enumerate(header_row):
                if not header:
                    continue
                record[str(header)] = row[index] if index < len(row) else ""
            data_rows.append(record)

        sheets.append(
            {
                "title": sheet.title,
                "rows": sheet.max_row,
                "cols": sheet.max_column,
                "header": header_row,
                "data": data_rows,
            }
        )

        if sheet.title == "分组统计":
            profile = _profile_from_stats_sheet(data_rows=data_rows, base_profile=profile)

    profile = _profile_from_storyboard_sheet(sheets=sheets, base_profile=profile)
    payload = {
        "source_workbook": str(workbook_path),
        "sheets": sheets,
        "profile": profile,
    }
    payload["profile"]["source"] = str(workbook_path)
    payload["profile"] = _normalize_storyboard_profile(payload["profile"], reference_payload=payload)
    REFERENCE_DIR.mkdir(parents=True, exist_ok=True)
    REFERENCE_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    PROFILE_JSON.write_text(json.dumps(payload["profile"], ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def _profile_from_stats_sheet(*, data_rows: list[dict[str, Any]], base_profile: dict[str, Any]) -> dict[str, Any]:
    profile = dict(base_profile)
    group_durations: list[float] = []
    for row in data_rows:
        group_name = str(row.get("分组", "")).strip()
        if not group_name or "总计" in group_name:
            continue
        duration = row.get("组时长(s)")
        if duration in ("", None):
            continue
        try:
            group_durations.append(float(duration))
        except (TypeError, ValueError):
            continue
    if group_durations:
        profile["group_durations"] = group_durations
        profile["group_count"] = len(group_durations)
        profile["target_duration_seconds"] = round(sum(group_durations), 2)
    return profile


def _profile_from_storyboard_sheet(*, sheets: list[dict[str, Any]], base_profile: dict[str, Any]) -> dict[str, Any]:
    profile = dict(base_profile)
    storyboard_sheet = next((sheet for sheet in sheets if "分镜" in str(sheet.get("title", ""))), None)
    if not storyboard_sheet:
        return _normalize_storyboard_profile(profile)

    rows = storyboard_sheet.get("data", [])
    if not rows:
        return _normalize_storyboard_profile(profile)

    grouped_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped_rows[str(row.get("分组", "")).strip()].append(row)

    style_blocks: list[dict[str, Any]] = []
    reference_group_shot_counts: list[int] = []
    for index, group_name in enumerate(sorted(grouped_rows.keys()), start=1):
        items = grouped_rows[group_name]
        reference_group_shot_counts.append(len(items))
        size_candidates = [item for item, _ in Counter(str(row.get("镜头景别", "")).strip() for row in items).most_common(4) if item]
        movement_candidates = [item for item, _ in Counter(str(row.get("镜头运动", "")).strip() for row in items).most_common(4) if item]
        dialogue_count = sum(1 for row in items if str(row.get("台词对白", "")).strip() not in {"", "—"})
        audio_count = sum(1 for row in items if str(row.get("音效", "")).strip() not in {"", "—"})
        default_block = DEFAULT_GROUP_STYLE_BLOCKS[min(index - 1, len(DEFAULT_GROUP_STYLE_BLOCKS) - 1)]
        style_blocks.append(
            {
                "group": index,
                "group_name": group_name,
                "beat": default_block["beat"],
                "focus": default_block["focus"],
                "size_candidates": size_candidates or list(default_block["size_candidates"]),
                "movement_candidates": movement_candidates or list(default_block["movement_candidates"]),
                "dialogue_density": "high" if dialogue_count >= max(1, len(items) // 2) else "medium",
                "audio_density": "high" if audio_count >= max(1, len(items) // 2) else "medium",
                "audio_focus": default_block["audio_focus"],
                "reference_shot_count": len(items),
            }
        )

    if style_blocks:
        profile["group_style_blocks"] = style_blocks
        profile["group_beats"] = [item["beat"] for item in style_blocks]
        profile["reference_group_shot_counts"] = reference_group_shot_counts
        profile["reference_total_shot_count"] = sum(reference_group_shot_counts)
        profile["default_group_shot_distribution"] = build_fallback_shot_distribution(
            group_durations=profile.get("group_durations", DEFAULT_STORYBOARD_PROFILE["group_durations"]),
            shot_count=int(profile.get("default_shot_count", 10) or 10),
        )
    return _normalize_storyboard_profile(profile)


def _normalize_storyboard_profile(profile: dict[str, Any], reference_payload: dict[str, Any] | None = None) -> dict[str, Any]:
    normalized = dict(DEFAULT_STORYBOARD_PROFILE)
    normalized.update(profile)
    normalized["group_beats"] = list(normalized.get("group_beats", DEFAULT_GROUP_BEATS))
    normalized["group_style_blocks"] = _normalize_group_style_blocks(normalized.get("group_style_blocks", []))
    normalized["default_group_shot_distribution"] = build_fallback_shot_distribution(
        group_durations=normalized.get("group_durations", DEFAULT_STORYBOARD_PROFILE["group_durations"]),
        shot_count=int(normalized.get("default_shot_count", 10) or 10),
    )
    if reference_payload and "source_workbook" in reference_payload:
        normalized["source"] = reference_payload["source_workbook"]
    return normalized


def _normalize_group_style_blocks(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not items:
        return [dict(item) for item in DEFAULT_GROUP_STYLE_BLOCKS]
    normalized: list[dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        default_item = DEFAULT_GROUP_STYLE_BLOCKS[min(index - 1, len(DEFAULT_GROUP_STYLE_BLOCKS) - 1)]
        payload = dict(default_item)
        payload.update(item)
        payload["size_candidates"] = list(payload.get("size_candidates", default_item["size_candidates"])) or list(default_item["size_candidates"])
        payload["movement_candidates"] = list(payload.get("movement_candidates", default_item["movement_candidates"])) or list(default_item["movement_candidates"])
        normalized.append(payload)
    return normalized


def build_fallback_shot_distribution(*, group_durations: list[float], shot_count: int) -> list[int]:
    group_count = len(group_durations)
    if group_count <= 0:
        return []
    if shot_count <= 0:
        return [0] * group_count
    distribution = [1] * group_count
    remaining = shot_count - group_count
    if remaining < 0:
        distribution = [0] * group_count
        for index in range(shot_count):
            distribution[index] = 1
        return distribution

    ranked_groups = sorted(range(group_count), key=lambda idx: group_durations[idx], reverse=True)
    cursor = 0
    while remaining > 0:
        distribution[ranked_groups[cursor % len(ranked_groups)]] += 1
        remaining -= 1
        cursor += 1
    return distribution


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        text = f"{value:.4f}".rstrip("0").rstrip(".")
        try:
            return float(text) if "." in text else int(text)
        except ValueError:
            return text
    return value
